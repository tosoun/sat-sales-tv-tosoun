import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

xar_data = [
    (211, "Δ.ΓΕΩΡΓΙΑΔΟΥ 24-ΛΑΡΙΣΑ"),
    (201, "ΝΙΚΗΤΑΡΑ 13 - ΛΑΡΙΣΑ"),
    (347, "ΙΩΑΝΝΙΝΩΝ 80 ΛΑΡΙΣΑ")import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

xar_data = [
    (211, "Δ.ΓΕΩΡΓΙΑΔΟΥ 24-ΛΑΡΙΣΑ"),
    (201, "ΝΙΚΗΤΑΡΑ 13 - ΛΑΡΙΣΑ"),
    (347, "ΙΩΑΝΝΙΝΩΝ 80 ΛΑΡΙΣΑ"),
    (212, "23ης ΟΚΤΩΒΡΙΟΥ 102-104 ΛΑΡΙΣΑ"),
    (239, "Χατζημιχάλη 49 - Φιλιππούπολη ΛΑΡΙΣΑ"),
    (219, "ΘΥΑΤΕΙΡΩΝ & Βενιζέλου Ν.Ιωνία - ΒΟΛΟΣ"),
    (220, "ΑΧΙΛΛΟΠΟΥΛΟΥ 171 - ΒΟΛΟΣ"),
    (222, "ΚΑΣΣΑΒΕΤΗ 14 & 28ης ΟΚΤΩΒΡΙΟΥ - ΒΟΛΟΣ"),
    (223, "Ν.Ιωνία - ΒΟΛΟΣ"),
    (240, "ΚΟΥΜΟΥΝΔΟΥΡΟΥ 150 - ΒΟΛΟΣ"),
    (242, "ΜΕΤΑΜΟΡΦΩΣΕΩΣ 21 & ΑΛΕΞΑΝΔΡΑΣ - ΒΟΛΟΣ"),
    (493, "ΣΚΙΑΘΟΣ")
]

mit_data = [
    (301, "ΑΙΑΝΗ - ΚΟΖΑΝΗ"),
    (302, "ΠΛ.ΕΛΕΥΘΕΡΙΑΣ - ΚΟΖΑΝΗ"),
    (309, "ΠΛ.ΛΑΣΣΑΝΗ 13 - ΚΟΖΑΝΗ"),
    (486, "Μ.ΑΛΕΞΑΝΔΡΟΥ 23 - ΣΙΑΤΙΣΤΑ"),
    (304, "25ης ΜΑΡΤΙΟΥ 74 - ΠΤΟΛΕΜΑΙΔΑ"),
    (352, "ΓΡΕΒΕΝΑ"),
    (308, "ΦΙΛΙΠΠΟΥ 1 - ΠΤΟΛΕΜΑΙΔΑ"),
    (353, "ΕΛ.ΒΕΝΙΖΕΛΟΥ & ΙΩΑΝΝΗ ΑΡΤΗ - ΦΛΩΡΙΝΑ"),
    (354, "ΚΟΖΑΝΗΣ ΚΑΙ ΓΡΕΒΕΝΩΝ - ΦΛΩΡΙΝΑ"),
    (355, "Χλόη Καστοριάς ΚΑΣΤΟΡΙΑ"),
    (356, "ΔΙΣΠΥΛΟ ΚΑΣΤΟΡΙΑΣ"),
    (366, "ΑΘ. ΔΙΑΚΟΥ 30 & ΓΡΑΜΜΟΥ 59"),
    (374, "ΜΑΝΙΑΚΟΙ ΚΑΣΤΟΡΙΑΣ")
]

pap_data = [
    (210, "ΕΛΑΣΣΩΝΑ ΒΥΖΑΝΤΙΟΥ"),
    (346, "ΛΑΡΙΣΗΣ 63 ΦΑΡΣΑΛΑ"),
    (202, "ΑΒΕΡΩΦ 22 - ΚΑΡΔΙΤΣΑ"),
    (204, "ΚΑΡΑΪΣΚΑΚΗ 95 - ΚΑΡΔΙΤΣΑ"),
    (206, "ΣΟΦΑΔΕΣ ΚΑΡΔΙΤΣΑΣ"),
    (209, "ΚΟΝΔΥΛΗ-ΤΡΙΚΑΛΑ"),
    (205, "ΚΟΝΔΥΛΗ 15 - ΤΡΙΚΑΛΑ"),
    (207, "ΔΕΛΗΓΙΩΡΓΗ-ΤΡΙΚΑΛΑ"),
    (208, "ΕΛΕΥΘΕΡΙΟΣ-ΤΡΙΚΑΛΑ"),
    (215, "ΑΒΕΡΩΦ ΚΑΙ ΛΑΡΙΣΗΣ")
]

pat_data = [
    (198, "ΓΚΟΥΡΑΣ & ΝΙΚΟΠΟΛΕΩΣ - ΙΩΑΝΝΙΝΑ"),
    (225, "Γ.ΠΑΠΑΝΔΡΕΟΥ 26-28 ΙΩΑΝΝΙΝΑ"),
    (226, "ΚΑΤΩ ΝΕΟΧΩΡΟΠΟΥΛΟ ΙΩΑΝΝΙΝΑ"),
    (316, "ΜΑΡΙΚΑΣ ΚΟΤΟΠΟΥΛΗ 66 – 68 ΙΩΑΝΝΙΝΑ ΡΙΖΑΡΙΟ"),
    (317, "ΛΕΩΦ. ΔΗΜΟΚΡΑΤΙΑΣ ΚΑΡΔΑΜΙΤΣΙΑ ΙΩΑΝΝΙΝΑ"),
    (381, "ΚΟΡΑΗ ΙΩΑΝΝΙΝΑ"),
    (227, "Μ. ΑΛΕΞΑΝΔΡΟΥ & Κ.ΠΑΛΑΙΟΛΟΓΟΥ - ΑΝΑΤΟΛΗ ΙΩΑΝΝΙΝΑ"),
    (228, "ΚΑΡΥΩΤΑΚΗ 15 & ΛΕΩΦ. ΕΙΡΗΝΗΣ - ΠΡΕΒΕΖΑ"),
    (229, "ΙΩΑΝΝΙΝΩΝ 199 - ΠΡΕΒΕΖΑ"),
    (224, "ΠΛ. ΚΙΛΚΙΣ & ΑΝΕΞΑΡΤΗΣΙΑΣ ΑΡΤΑ"),
    (315, "26ο χλμ ΕΘΝΙΚΗΣ ΟΔΟΥ ΠΡΕΒΕΖΗΣ - ΙΩΑΝΝΙΝΩΝ ΛΟΥΡΟΣ"),
    (359, "ΗΓΟΥΜΕΝΙΤΣΑ ΘΕΣΠΡΩΤΙΑΣ"),
    (378, "ΠΑΡΑΜΥΘΙΑ ΘΕΣΠΡΩΤΙΑΣ")
]

ski_data = [
    (531, "ΚΕΡΚΥΡΑ ΑΧΑΡΑΒΗ"),
    (539, "ΚΕΡΚΥΡΑ ΚΑΣΣΙΩΠΗ"),
    (567, "ΚΕΡΚΥΡΑ ΣΙΔΑΡΙ ΚΑΡΟΥΣΑΔΕΣ ΜΑΡΚΑΤΟ"),
    (537, "ΚΕΡΚΥΡΑ ΜΑΝΤΟΥΚΙ"),
    (533, "ΚΕΡΚΥΡΑ ΑΛΥΚΕΣ"),
    (525, "ΚΕΡΚΥΡΑ ΥΠΕΡ ΕΘΝΙΚΗ ΟΔΟΣ ΛΕΥΚΙΜΜΗΣ ΚΕΡΚΥΡΑΣ"),
    (535, "ΚΕΡΚΥΡΑ ΛΕΥΚΙΜΜΗ"),
    (534, "ΚΕΡΚΥΡΑ ΜΩΡΑΙΤΙΚΑ"),
    (566, "ΚΕΡΚΥΡΑ ΚΟΜΒΟΣ ΒΡΥΩΝΗ ΚΑΣΤΕΛΛΑΝΟΙ ΜΑΡΚΑΤΟ"),
    (540, "ΚΕΡΚΥΡΑ ΑΛΕΠΟΥ ΜΑΡΚΑΤΟ"),
    (565, "ΚΕΡΚΥΡΑ ΣΑΡΟΚΟ ΜΑΡΚΑΤΟ"),
    (530, "ΚΕΡΚΥΡΑ ΙΩΑΝΝΟΥ ΘΕΟΤΟΚΗ ΠΑΛΛΑΔΑ"),
    (528, "ΚΕΡΚΥΡΑ ΛΑΙΚΗ ΑΓΟΡΑ ΓΕΡΑΣΙΜΟΥ ΜΑΡΚΟΡΑ"),
    (529, "ΚΕΡΚΥΡΑ ΠΙΝΙΑ"),
    (532, "ΚΕΡΚΥΡΑ ΝΟΣΟΚΟΜΕΙΟ"),
    (536, "ΚΕΡΚΥΡΑ ΣΠΗΛΙΑ"),
    (538, "ΚΕΡΚΥΡΑ ΜΗΤΡΟΠΟΛΙΤΟΥ ΜΕΘΟΔΙΟΥ")
]

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

ws_summary = wb.create_sheet(title="ΕΥΡΕΤΗΡΙΟ & SEARCH")
ws_summary.views.sheetView[0].showGridLines = True

ws_summary['A1'] = "ΣΥΣΤΗΜΑ ΟΜΑΔΩΝ ΚΑΤΑΣΤΗΜΑΤΩΝ"
ws_summary['A1'].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")

headers_summary = ["ΓΚΡΟΥΠ", "ΚΩΔΙΚΟΣ", "ΔΙΕΥΘΥΝΣΗ / ΠΟΛΗ", "ΚΑΤΑΣΤΑΣΗ"]
for col_idx, h in enumerate(headers_summary, start=1):
    cell = ws_summary.cell(row=3, column=col_idx, value=h)
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

row_idx = 4
all_groups = [("ΧΑΡ", xar_data), ("ΜΗΤ", mit_data), ("ΠΑΠ", pap_data), ("ΠΑΤ", pat_data), ("ΣΚΙ", ski_data)]

for group_name, group_data in all_groups:
    for code, name in group_data:
        ws_summary.cell(row=row_idx, column=1, value=group_name).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=2, value=code).alignment = Alignment(horizontal="center")
        ws_summary.cell(row=row_idx, column=3, value=name)
        ws_summary.cell(row=row_idx, column=4, value="Ενεργό").alignment = Alignment(horizontal="center")
        for c in range(1, 5): 
            ws_summary.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1

def create_group_sheet(ws_name, title_text, data, header_color, fill_color, text_color):
    ws = wb.create_sheet(title=ws_name)
    ws.views.sheetView[0].showGridLines = True
    ws['A1'] = title_text
    ws['A1'].font = Font(name="Calibri", size=14, bold=True, color=text_color)

    for col_idx, h in enumerate(["ΚΩΔΙΚΟΣ", "ΔΙΕΥΘΥΝΣΗ / ΠΟΛΗ"], start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r = 4
    for code, name in data:
        c1 = ws.cell(row=r, column=1, value=code)
        c1.alignment = Alignment(horizontal="center")
        c1.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        c2 = ws.cell(row=r, column=2, value=name)
        c2.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for c in range(1, 3): 
            ws.cell(row=r, column=c).border = thin_border
        r += 1

create_group_sheet("ΓΚΡΟΥΠ ΧΑΡ", "ΚΑΤΑΣΤΗΜΑΤΑ ΧΑΡ", xar_data, "38761D", "D9EAD3", "274E13")
create_group_sheet("ΓΚΡΟΥΠ ΜΗΤ", "ΚΑΤΑΣΤΗΜΑΤΑ ΜΗΤ", mit_data, "B45F06", "FFF2CC", "7F6000")
create_group_sheet("ΓΚΡΟΥΠ ΠΑΠ", "ΚΑΤΑΣΤΗΜΑΤΑ ΠΑΠ", pap_data, "00838F", "E0F7FA", "134F5C")
create_group_sheet("ΓΚΡΟΥΠ ΠΑΤ", "ΚΑΤΑΣΤΗΜΑΤΑ ΠΑΤ", pat_data, "B8860B", "FEF9E7", "7E5109")
create_group_sheet("ΓΚΡΟΥΠ ΣΚΙ", "ΚΑΤΑΣΤΗΜΑΤΑ ΣΚΙ", ski_data, "D35400", "FADBD8", "78281F")

for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 2:
                val_str = str(cell.value or '')
                if len(val_str) > max_len: max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

wb.save("Katastimata_Clean.xlsx")
print("Saved!"),
    (212, "23ης ΟΚΤΩΒΡΙΟΥ 102-104 ΛΑΡΙΣΑ"),
    (239, "Χατζημιχάλη 49 - Φιλιππούπολη ΛΑΡΙΣΑ"),
    (219, "ΘΥΑΤΕΙΡΩΝ & Βενιζέλου Ν.Ιωνία - ΒΟΛΟΣ"),
    (220, "ΑΧΙΛΛΟΠΟΥΛΟΥ 171 - ΒΟΛΟΣ"),
    (222, "ΚΑΣΣΑΒΕΤΗ 14 & 28ης ΟΚΤΩΒΡΙΟΥ - ΒΟΛΟΣ"),
    (223, "Ν.Ιωνία - ΒΟΛΟΣ"),
    (240, "ΚΟΥΜΟΥΝΔΟΥΡΟΥ 150 - ΒΟΛΟΣ"),
    (242, "ΜΕΤΑΜΟΡΦΩΣΕΩΣ 21 & ΑΛΕΞΑΝΔΡΑΣ - ΒΟΛΟΣ"),
    (493, "ΣΚΙΑΘΟΣ")
]

mit_data = [
    (301, "ΑΙΑΝΗ - ΚΟΖΑΝΗ"),
    (302, "ΠΛ.ΕΛΕΥΘΕΡΙΑΣ - ΚΟΖΑΝΗ"),
    (309, "ΠΛ.ΛΑΣΣΑΝΗ 13 - ΚΟΖΑΝΗ"),
    (486, "Μ.ΑΛΕΞΑΝΔΡΟΥ 23 - ΣΙΑΤΙΣΤΑ"),
    (304, "25ης ΜΑΡΤΙΟΥ 74 - ΠΤΟΛΕΜΑΙΔΑ"),
    (352, "ΓΡΕΒΕΝΑ"),
    (308, "ΦΙΛΙΠΠΟΥ 1 - ΠΤΟΛΕΜΑΙΔΑ"),
    (353, "ΕΛ.ΒΕΝΙΖΕΛΟΥ & ΙΩΑΝΝΗ ΑΡΤΗ - ΦΛΩΡΙΝΑ"),
    (354, "ΚΟΖΑΝΗΣ ΚΑΙ ΓΡΕΒΕΝΩΝ - ΦΛΩΡΙΝΑ"),
    (355, "Χλόη Καστοριάς ΚΑΣΤΟΡΙΑ"),
    (356, "ΔΙΣΠΥΛΟ ΚΑΣΤΟΡΙΑΣ"),
    (366, "ΑΘ. ΔΙΑΚΟΥ 30 & ΓΡΑΜΜΟΥ 59"),
    (374, "ΜΑΝΙΑΚΟΙ ΚΑΣΤΟΡΙΑΣ")
]

pap_data = [
    (210, "ΕΛΑΣΣΩΝΑ ΒΥΖΑΝΤΙΟΥ"),
    (346, "ΛΑΡΙΣΗΣ 63 ΦΑΡΣΑΛΑ"),
    (202, "ΑΒΕΡΩΦ 22 - ΚΑΡΔΙΤΣΑ"),
    (204, "ΚΑΡΑΪΣΚΑΚΗ 95 - ΚΑΡΔΙΤΣΑ"),
    (206, "ΣΟΦΑΔΕΣ ΚΑΡΔΙΤΣΑΣ"),
    (209, "ΚΟΝΔΥΛΗ-ΤΡΙΚΑΛΑ"),
    (205, "ΚΟΝΔΥΛΗ 15 - ΤΡΙΚΑΛΑ"),
    (207, "ΔΕΛΗΓΙΩΡΓΗ-ΤΡΙΚΑΛΑ"),
    (208, "ΕΛΕΥΘΕΡΙΟΣ-ΤΡΙΚΑΛΑ"),
    (215, "ΑΒΕΡΩΦ ΚΑΙ ΛΑΡΙΣΗΣ")
]

pat_data = [
    (198, "ΓΚΟΥΡΑΣ & ΝΙΚΟΠΟΛΕΩΣ - ΙΩΑΝΝΙΝΑ"),
    (225, "Γ.ΠΑΠΑΝΔΡΕΟΥ 26-28 ΙΩΑΝΝΙΝΑ"),
    (226, "ΚΑΤΩ ΝΕΟΧΩΡΟΠΟΥΛΟ ΙΩΑΝΝΙΝΑ"),
    (316, "ΜΑΡΙΚΑΣ ΚΟΤΟΠΟΥΛΗ 66 – 68 ΙΩΑΝΝΙΝΑ ΡΙΖΑΡΙΟ"),
    (317, "ΛΕΩΦ. ΔΗΜΟΚΡΑΤΙΑΣ ΚΑΡΔΑΜΙΤΣΙΑ ΙΩΑΝΝΙΝΑ"),
    (381, "ΚΟΡΑΗ ΙΩΑΝΝΙΝΑ"),
    (227, "Μ. ΑΛΕΞΑΝΔΡΟΥ & Κ.ΠΑΛΑΙΟΛΟΓΟΥ - ΑΝΑΤΟΛΗ ΙΩΑΝΝΙΝΑ"),
    (228, "ΚΑΡΥΩΤΑΚΗ 15 & ΛΕΩΦ. ΕΙΡΗΝΗΣ - ΠΡΕΒΕΖΑ"),
    (229, "ΙΩΑΝΝΙΝΩΝ 199 - ΠΡΕΒΕΖΑ"),
    (224, "ΠΛ. ΚΙΛΚΙΣ & ΑΝΕΞΑΡΤΗΣΙΑΣ ΑΡΤΑ"),
    (315, "26ο χλμ ΕΘΝΙΚΗΣ ΟΔΟΥ ΠΡΕΒΕΖΗΣ - ΙΩΑΝΝΙΝΩΝ ΛΟΥΡΟΣ"),
    (359, "ΗΓΟΥΜΕΝΙΤΣΑ ΘΕΣΠΡΩΤΙΑΣ"),
    (378, "ΠΑΡΑΜΥΘΙΑ ΘΕΣΠΡΩΤΙΑΣ")
]

ski_data = [
    (531, "ΚΕΡΚΥΡΑ ΑΧΑΡΑΒΗ"),
    (539, "ΚΕΡΚΥΡΑ ΚΑΣΣΙΩΠΗ"),
    (567, "ΚΕΡΚΥΡΑ ΣΙΔΑΡΙ ΚΑΡΟΥΣΑΔΕΣ ΜΑΡΚΑΤΟ"),
    (537, "ΚΕΡΚΥΡΑ ΜΑΝΤΟΥΚΙ"),
    (533, "ΚΕΡΚΥΡΑ ΑΛΥΚΕΣ"),
    (525, "ΚΕΡΚΥΡΑ ΥΠΕΡ ΕΘΝΙΚΗ ΟΔΟΣ ΛΕΥΚΙΜΜΗΣ ΚΕΡΚΥΡΑΣ"),
    (535, "ΚΕΡΚΥΡΑ ΛΕΥΚΙΜΜΗ"),
    (534, "ΚΕΡΚΥΡΑ ΜΩΡΑΙΤΙΚΑ"),
    (566, "ΚΕΡΚΥΡΑ ΚΟΜΒΟΣ ΒΡΥΩΝΗ ΚΑΣΤΕΛΛΑΝΟΙ ΜΑΡΚΑΤΟ"),
    (540, "ΚΕΡΚΥΡΑ ΑΛΕΠΟΥ ΜΑΡΚΑΤΟ"),
    (565, "ΚΕΡΚΥΡΑ ΣΑΡΟΚΟ ΜΑΡΚΑΤΟ"),
    (530, "ΚΕΡΚΥΡΑ ΙΩΑΝΝΟΥ ΘΕΟΤΟΚΗ ΠΑΛΛΑΔΑ"),
    (528, "ΚΕΡΚΥΡΑ ΛΑΙΚΗ ΑΓΟΡΑ ΓΕΡΑΣΙΜΟΥ ΜΑΡΚΟΡΑ"),
    (529, "ΚΕΡΚΥΡΑ ΠΙΝΙΑ"),
    (532, "ΚΕΡΚΥΡΑ ΝΟΣΟΚΟΜΕΙΟ"),
    (536, "ΚΕΡΚΥΡΑ ΣΠΗΛΙΑ"),
    (538, "ΚΕΡΚΥΡΑ ΜΗΤΡΟΠΟΛΙΤΟΥ ΜΕΘΟΔΙΟΥ")
]

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

ws_summary = wb.create_sheet(title="ΕΥΡΕΤΗΡΙΟ & SEARCH")
ws_summary.views.sheetView[0].showGridLines = True

ws_summary['A1'] = "ΣΥΣΤΗΜΑ ΟΜΑΔΩΝ ΚΑΤΑΣΤΗΜΑΤΩΝ & ΑΝΑΖΗΤΗΣΗΣ"
ws_summary['A1'].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
ws_summary.merge_cells('A1:E1')

ws_summary['A3'] = "Οδηγίες: Πληκτρολογήστε το Γκρουπ (ΧΑΡ, ΜΗΤ, ΠΑΠ, ΠΑΤ, ΣΚΙ) στο φίλτρο ή επιλέξτε καρτέλα."
ws_summary['A3'].font = Font(name="Calibri", size=11, italic=True)

headers_summary = ["ΓΚΡΟΥΠ", "ΚΩΔΙΚΟΣ", "ΔΙΕΥΘΥΝΣΗ / ΠΟΛΗ", "ΠΩΛΗΣΕΙΣ (ΕΥΡΩ)", "ΚΑΤΑΣΤΑΣΗ"]
for col_idx, h in enumerate(headers_summary, start=1):
    cell = ws_summary.cell(row=5, column=col_idx, value=h)
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

row_idx = 6
for code, name in xar_data:
    ws_summary.cell(row=row_idx, column=1, value="ΧΑΡ").alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=2, value=code).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=3, value=name)
    ws_summary.cell(row=row_idx, column=4, value="=RANDBETWEEN(15000, 45000)").number_format = '#,##0.00 €'
    ws_summary.cell(row=row_idx, column=5, value="Ενεργό").alignment = Alignment(horizontal="center")
    for c in range(1, 6): ws_summary.cell(row=row_idx, column=c).border = thin_border
    row_idx += 1

for code, name in mit_data:
    ws_summary.cell(row=row_idx, column=1, value="ΜΗΤ").alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=2, value=code).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=3, value=name)
    ws_summary.cell(row=row_idx, column=4, value="=RANDBETWEEN(15000, 45000)").number_format = '#,##0.00 €'
    ws_summary.cell(row=row_idx, column=5, value="Ενεργό").alignment = Alignment(horizontal="center")
    for c in range(1, 6): ws_summary.cell(row=row_idx, column=c).border = thin_border
    row_idx += 1

for code, name in pap_data:
    ws_summary.cell(row=row_idx, column=1, value="ΠΑΠ").alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=2, value=code).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=3, value=name)
    ws_summary.cell(row=row_idx, column=4, value="=RANDBETWEEN(15000, 45000)").number_format = '#,##0.00 €'
    ws_summary.cell(row=row_idx, column=5, value="Ενεργό").alignment = Alignment(horizontal="center")
    for c in range(1, 6): ws_summary.cell(row=row_idx, column=c).border = thin_border
    row_idx += 1

for code, name in pat_data:
    ws_summary.cell(row=row_idx, column=1, value="ΠΑΤ").alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=2, value=code).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=3, value=name)
    ws_summary.cell(row=row_idx, column=4, value="=RANDBETWEEN(15000, 45000)").number_format = '#,##0.00 €'
    ws_summary.cell(row=row_idx, column=5, value="Ενεργό").alignment = Alignment(horizontal="center")
    for c in range(1, 6): ws_summary.cell(row=row_idx, column=c).border = thin_border
    row_idx += 1

for code, name in ski_data:
    ws_summary.cell(row=row_idx, column=1, value="ΣΚΙ").alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=2, value=code).alignment = Alignment(horizontal="center")
    ws_summary.cell(row=row_idx, column=3, value=name)
    ws_summary.cell(row=row_idx, column=4, value="=RANDBETWEEN(15000, 45000)").number_format = '#,##0.00 €'
    ws_summary.cell(row=row_idx, column=5, value="Ενεργό").alignment = Alignment(horizontal="center")
    for c in range(1, 6): ws_summary.cell(row=row_idx, column=c).border = thin_border
    row_idx += 1

ws_summary.auto_filter.ref = f"A5:E{row_idx-1}"

def create_group_sheet(ws_name, title_text, data, header_color, fill_color, text_color):
    ws = wb.create_sheet(title=ws_name)
    ws.views.sheetView[0].showGridLines = True
    ws['A1'] = title_text
    ws['A1'].font = Font(name="Calibri", size=14, bold=True, color=text_color)
    ws.merge_cells('A1:C1')

    for col_idx, h in enumerate(["ΚΩΔΙΚΟΣ", "ΔΙΕΥΘΥΝΣΗ / ΠΟΛΗ", "ΠΩΛΗΣΕΙΣ"], start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r = 4
    for code, name in data:
        c1 = ws.cell(row=r, column=1, value=code)
        c1.alignment = Alignment(horizontal="center")
        c1.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        c2 = ws.cell(row=r, column=2, value=name)
        c2.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        c3 = ws.cell(row=r, column=3, value=f"=VLOOKUP(A{r}, 'ΕΥΡΕΤΗΡΙΟ & SEARCH'!B:D, 3, FALSE)")
        c3.number_format = '#,##0.00 €'
        c3.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for c in range(1, 4): ws.cell(row=r, column=c).border = thin_border
        r += 1
    return ws

create_group_sheet("ΓΚΡΟΥΠ ΧΑΡ", "ΚΑΤΑΣΤΗΜΑΤΑ ΓΚΡΟΥΠ ΧΑΡ", xar_data, "38761D", "D9EAD3", "274E13")
create_group_sheet("ΓΚΡΟΥΠ ΜΗΤ", "ΚΑΤΑΣΤΗΜΑΤΑ ΓΚΡΟΥΠ ΜΗΤ", mit_data, "B45F06", "FFF2CC", "7F6000")
create_group_sheet("ΓΚΡΟΥΠ ΠΑΠ", "ΚΑΤΑΣΤΗΜΑΤΑ ΓΚΡΟΥΠ ΠΑΠ", pap_data, "00838F", "E0F7FA", "134F5C")
create_group_sheet("ΓΚΡΟΥΠ ΠΑΤ", "ΚΑΤΑΣΤΗΜΑΤΑ ΓΚΡΟΥΠ ΠΑΤ", pat_data, "B8860B", "FEF9E7", "7E5109")
create_group_sheet("ΓΚΡΟΥΠ ΣΚΙ", "ΚΑΤΑΣΤΗΜΑΤΑ ΓΚΡΟΥΠ ΣΚΙ", ski_data, "D35400", "FADBD8", "78281F")

for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 2:
                val_str = str(cell.value or '')
                if len(val_str) > max_len: max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

wb.save("Katastimata_All_Groups_Complete.xlsx")
